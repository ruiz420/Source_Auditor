from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import csv
import hmac
import os
import re
import requests
import time
import uuid
from datetime import datetime
from functools import wraps
from io import BytesIO, StringIO
from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-before-production")

USER_STATES = {}
REQUEST_TIMEOUT = 8
USER_AGENTS = {
    "chrome": {
        "label": "Google Chrome",
        "user_agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
    },
    "firefox": {
        "label": "Mozilla Firefox",
        "user_agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) "
            "Gecko/20100101 Firefox/125.0"
        ),
    },
    "googlebot": {
        "label": "Googlebot",
        "user_agent": "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)",
    },
    "mobile": {
        "label": "Mobile Safari",
        "user_agent": (
            "Mozilla/5.0 (iPhone; CPU iPhone OS 17_4 like Mac OS X) "
            "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Mobile/15E148 Safari/604.1"
        ),
    },
}
DEFAULT_AGENT = "chrome"


def build_headers(agent_key):
    user_agent = USER_AGENTS.get(agent_key, USER_AGENTS[DEFAULT_AGENT])["user_agent"]

    return {
        "User-Agent": user_agent,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8,pt-BR;q=0.7,pt;q=0.6",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    }


BROWSER_HEADERS = build_headers(DEFAULT_AGENT)
AGENT_ORDER = ["chrome", "firefox", "googlebot", "mobile"]
SCREENSHOT_DIR = os.path.join(app.root_path, "static", "screenshots")

LEGACY_BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8,pt-BR;q=0.7,pt;q=0.6",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}


def load_allowed_users():
    users_file = os.environ.get("AUDITOR_USERS_FILE", os.path.join(app.root_path, "usuarios.txt"))
    if os.path.exists(users_file):
        users = {}
        with open(users_file, "r", encoding="utf-8") as file:
            for line in file:
                clean_line = line.strip()
                if not clean_line or clean_line.startswith("#") or ":" not in clean_line:
                    continue
                username, password = clean_line.split(":", 1)
                username = username.strip()
                if username:
                    users[username] = password.strip()
        if users:
            return users

    raw_users = os.environ.get("AUDITOR_USERS", "").strip()
    if raw_users:
        users = {}
        for item in raw_users.split(","):
            if ":" not in item:
                continue
            username, password = item.split(":", 1)
            username = username.strip()
            if username:
                users[username] = password.strip()
        if users:
            return users

    username = os.environ.get("AUDITOR_USERNAME", "admin")
    password = os.environ.get("AUDITOR_PASSWORD", "admin123")
    return {username: password}


def is_valid_user(username, password):
    expected_password = load_allowed_users().get(username)
    return expected_password is not None and hmac.compare_digest(expected_password, password)


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if session.get("authenticated"):
            return view(*args, **kwargs)

        if request.path.startswith(("/add", "/check", "/mark", "/capture")):
            return jsonify({"error": "sesion no autorizada"}), 401

        return redirect(url_for("login", next=request.path))

    return wrapped_view


def get_state():
    session_id = session.get("session_id")
    if not session_id:
        session_id = str(uuid.uuid4())
        session["session_id"] = session_id

    return USER_STATES.setdefault(session_id, {"urls": [], "results": {}})


def fix_url(url):
    url = url.strip()
    if not url.lower().startswith(("http://", "https://")):
        return "https://" + url
    return url


def extract_title(html):
    match = re.search(r"<title[^>]*>(.*?)</title>", html, re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    return re.sub(r"\s+", " ", match.group(1)).strip()


def visible_text_length(html):
    without_scripts = re.sub(
        r"<(script|style)[^>]*>.*?</\1>",
        "",
        html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    text = re.sub(r"<[^>]+>", " ", without_scripts)
    text = re.sub(r"\s+", " ", text).strip()
    return len(text)


def analyze_response(response):
    html = response.text or ""
    lowered = html[:30000].lower()
    title = extract_title(html)
    text_length = visible_text_length(html)

    error_signals = [
        "404 not found",
        "page not found",
        "server error",
        "internal server error",
        "bad gateway",
        "service unavailable",
        "domain for sale",
        "parkingcrew",
        "this domain is parked",
    ]

    protection_signals = [
        "checking your browser",
        "just a moment",
        "cf-browser-verification",
        "access denied",
    ]

    is_blocking_code = response.status_code in (401, 403, 429, 503)
    has_browser_challenge = any(signal in lowered for signal in protection_signals)

    if response.status_code >= 400:
        if is_blocking_code and has_browser_challenge:
            return "DUDOSA", f"proteccion anti-bot con HTTP {response.status_code}", title, text_length
        return "ERROR", f"codigo HTTP {response.status_code}", title, text_length

    if has_browser_challenge and (text_length < 500 or not title):
        return "DUDOSA", "posible proteccion anti-bot", title, text_length

    for signal in error_signals:
        if signal in lowered:
            return "ERROR", f"contenido sospechoso: {signal}", title, text_length

    if text_length < 80 and not title:
        return "ERROR", "contenido vacio o sin titulo", title, text_length

    if text_length < 80:
        return "DUDOSA", "contenido muy corto", title, text_length

    if not title:
        return "DUDOSA", "pagina sin titulo HTML", title, text_length

    if response.history:
        return "VALIDA", "redireccion correcta", title, text_length

    return "VALIDA", "pagina responde correctamente", title, text_length


def is_good_result(data):
    return data["status"] == "VALIDA" and data.get("http_code", 0) == 200


def get_redirect_chain(url, agent_key=DEFAULT_AGENT):
    chain = []
    start = time.time()

    try:
        response = requests.get(
            url,
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,
            headers=build_headers(agent_key),
        )

        for history_item in response.history:
            chain.append(history_item.url)

        chain.append(response.url)
        load_time = round(time.time() - start, 2)
        status, reason, title, text_length = analyze_response(response)

        return {
            "final": response.url,
            "chain": chain,
            "status": status,
            "reason": reason,
            "http_code": response.status_code,
            "title": title,
            "text_length": text_length,
            "time": load_time,
            "agent": agent_key,
            "agent_label": USER_AGENTS.get(agent_key, USER_AGENTS[DEFAULT_AGENT])["label"],
            "screenshot": "",
        }

    except requests.Timeout:
        return {
            "final": "",
            "chain": [],
            "status": "ERROR",
            "reason": "timeout al conectar",
            "http_code": "",
            "title": "",
            "text_length": 0,
            "time": 0,
            "agent": agent_key,
            "agent_label": USER_AGENTS.get(agent_key, USER_AGENTS[DEFAULT_AGENT])["label"],
            "screenshot": "",
        }

    except requests.RequestException as exc:
        return {
            "final": "",
            "chain": [],
            "status": "ERROR",
            "reason": f"error de conexion: {exc.__class__.__name__}",
            "http_code": "",
            "title": "",
            "text_length": 0,
            "time": 0,
            "agent": agent_key,
            "agent_label": USER_AGENTS.get(agent_key, USER_AGENTS[DEFAULT_AGENT])["label"],
            "screenshot": "",
        }


def get_redirect_chain_auto(url, requested_agent=DEFAULT_AGENT):
    if requested_agent != "auto":
        return get_redirect_chain(url, requested_agent)

    best = None
    for agent_key in AGENT_ORDER:
        data = get_redirect_chain(url, agent_key)

        if is_good_result(data):
            return data

        if best is None:
            best = data
            continue

        if data["status"] == "VALIDA" and best["status"] != "VALIDA":
            best = data
        elif data["status"] == "DUDOSA" and best["status"] == "ERROR":
            best = data
        elif data.get("http_code") == 200 and best.get("http_code") != 200:
            best = data

    if best:
        best["reason"] = f"{best['reason']} (probado con varios agentes)"
        best["agent"] = "auto/" + best.get("agent", "")
        best["agent_label"] = "Auto: " + best.get("agent_label", "")
    return best


def safe_filename(value):
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "_", value.strip())[:80]
    return cleaned or "captura"


def capture_page_screenshot(url, original_url):
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return "Playwright no instalado"

    os.makedirs(SCREENSHOT_DIR, exist_ok=True)
    filename = f"{safe_filename(original_url)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    path = os.path.join(SCREENSHOT_DIR, filename)

    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 1366, "height": 768})
            page.goto(url, wait_until="networkidle", timeout=20000)
            page.screenshot(path=path, full_page=True)
            browser.close()
        return f"/static/screenshots/{filename}"
    except Exception as exc:
        return f"captura fallida: {exc.__class__.__name__}"


@app.route("/")
@login_required
def home():
    return render_template("dashboard.html")


@app.route("/auditor")
@login_required
def auditor():
    return render_template("index.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        if is_valid_user(username, password):
            session.clear()
            session["authenticated"] = True
            session["username"] = username
            session["session_id"] = str(uuid.uuid4())
            return redirect(request.args.get("next") or url_for("auditor"))

        error = "Usuario o clave incorrectos."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session_id = session.get("session_id")
    if session_id:
        USER_STATES.pop(session_id, None)
    session.clear()
    return redirect(url_for("login"))


@app.route("/add", methods=["POST"])
@login_required
def add():
    state = get_state()

    data = request.json["urls"]
    agent = request.json.get("agent", "auto")
    capture = bool(request.json.get("capture", False))
    state["results"] = {}
    state["urls"] = [
        {
            "original": raw_url.strip(),
            "request_url": fix_url(raw_url),
            "agent": agent,
            "capture": capture,
        }
        for raw_url in data.split("\n")
        if raw_url.strip()
    ]

    return jsonify({
        "total": len(state["urls"]),
        "urls": [item["original"] for item in state["urls"]],
    })


@app.route("/check")
@login_required
def check():
    state = get_state()
    output = []

    for item in state["urls"]:
        output.append(audit_item(item))

    return jsonify(output)


def audit_item(item):
    state = get_state()
    results = state["results"]
    original_url = item["original"]
    request_url = item["request_url"]

    data = get_redirect_chain_auto(request_url, item.get("agent", "auto"))
    data["original"] = original_url
    data["request_url"] = request_url
    data["redirected"] = bool(data["final"] and data["final"].rstrip("/") != request_url.rstrip("/"))

    if item.get("capture"):
        data["screenshot"] = capture_page_screenshot(data["final"] or request_url, original_url)

    results[original_url] = data

    return {
        "url": original_url,
        "request_url": request_url,
        "final_url": data["final"],
        "redirected": data["redirected"],
        "status": data["status"],
        "reason": data["reason"],
        "http_code": data["http_code"],
        "title": data["title"],
        "text_length": data["text_length"],
        "time": data["time"],
        "chain": data["chain"],
        "agent": data.get("agent", ""),
        "agent_label": data.get("agent_label", ""),
        "screenshot": data.get("screenshot", ""),
    }


@app.route("/check-one", methods=["POST"])
@login_required
def check_one():
    state = get_state()
    urls = state["urls"]
    index = int(request.json["index"])

    if index < 0 or index >= len(urls):
        return jsonify({"error": "indice fuera de rango"}), 400

    return jsonify(audit_item(urls[index]))


@app.route("/mark", methods=["POST"])
@login_required
def mark():
    state = get_state()
    results = state["results"]
    data = request.json

    url = data["url"]
    status = data["status"]

    if url not in results:
        results[url] = {}

    results[url]["manual"] = status

    return jsonify({"ok": True})


@app.route("/capture-one-manual", methods=["POST"])
@login_required
def capture_one_manual():
    state = get_state()
    urls = state["urls"]
    results = state["results"]
    index = int(request.json["index"])
    if index < 0 or index >= len(urls):
        return jsonify({"error": "indice fuera de rango"}), 400

    item = urls[index]
    original_url = item["original"]
    request_url = item["request_url"]

    # Obtenemos la URL final del resultado, o la URL de petición si no está
    final_url = results.get(original_url, {}).get("final") or request_url

    # Ejecutamos la captura con Playwright
    screenshot_path = capture_page_screenshot(final_url, original_url)

    # Actualizamos en memoria
    if original_url in results:
        results[original_url]["screenshot"] = screenshot_path
    else:
        results[original_url] = {"screenshot": screenshot_path}
    urls[index]["screenshot"] = screenshot_path

    return jsonify({
        "success": not screenshot_path.startswith("captura fallida"),
        "screenshot": screenshot_path
    })


@app.route("/export")
@login_required
def export():
    results = get_state()["results"]
    text_output = StringIO()
    writer = csv.writer(text_output)

    writer.writerow([
        "url",
        "resultado",
        "redireccion",
        "motivo",
        "codigo_http",
        "titulo",
        "texto_detectado",
        "tiempo_carga",
        "agente_usuario",
        "captura",
        "cadena_redireccion",
    ])

    for url, data in results.items():
        writer.writerow([
            url,
            data.get("manual") or data.get("status", ""),
            data.get("final", ""),
            data.get("reason", ""),
            data.get("http_code", ""),
            data.get("title", ""),
            data.get("text_length", ""),
            data.get("time", ""),
            data.get("agent_label", ""),
            data.get("screenshot", ""),
            " -> ".join(data.get("chain", [])),
        ])

    output = BytesIO()
    output.write(("\ufeff" + text_output.getvalue()).encode("utf-8"))
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="auditoria_final.csv",
        mimetype="text/csv; charset=utf-8",
    )


@app.route("/export-excel")
@login_required
def export_excel():
    results = get_state()["results"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria de Fuentes"

    # Estilos
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="0F2742", end_color="0F2742", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Colores por estado
    fill_valida = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_valida = Font(name=font_family, size=10, color="15803D", bold=True)

    fill_dudosa = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_dudosa = Font(name=font_family, size=10, color="A15C00", bold=True)

    fill_error = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_error = Font(name=font_family, size=10, color="C83232", bold=True)

    headers = [
        "URL Original",
        "Resultado",
        "Redireccion",
        "Motivo",
        "Codigo HTTP",
        "Titulo",
        "Tiempo Carga (s)",
        "Agente Usuario",
        "Captura de Pantalla",
        "Cadena Redireccion"
    ]

    # Escribir cabeceras
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    ws.row_dimensions[1].height = 28

    # Recorrer resultados
    for r_idx, (url, data) in enumerate(results.items(), start=2):
        status = data.get("manual") or data.get("status", "")
        final_url = data.get("final", "")
        reason = data.get("reason", "")
        http_code = data.get("http_code", "")
        title = data.get("title", "")
        time_load = data.get("time", "")
        agent = data.get("agent_label", "")
        chain = " -> ".join(data.get("chain", []))

        ws.cell(row=r_idx, column=1, value=url).alignment = align_left
        ws.cell(row=r_idx, column=2, value=status).alignment = align_center
        ws.cell(row=r_idx, column=3, value=final_url).alignment = align_left
        ws.cell(row=r_idx, column=4, value=reason).alignment = align_left
        ws.cell(row=r_idx, column=5, value=http_code).alignment = align_center
        ws.cell(row=r_idx, column=6, value=title).alignment = align_left
        ws.cell(row=r_idx, column=7, value=time_load).alignment = align_center
        ws.cell(row=r_idx, column=8, value=agent).alignment = align_left
        ws.cell(row=r_idx, column=10, value=chain).alignment = align_left

        # Aplicar fuente
        for col_idx in range(1, 11):
            if col_idx != 9:
                ws.cell(row=r_idx, column=col_idx).font = data_font

        # Estilo del estado
        cell_status = ws.cell(row=r_idx, column=2)
        status_upper = str(status).upper()
        if "VALIDA" in status_upper:
            cell_status.fill = fill_valida
            cell_status.font = font_valida
        elif "DUDOSA" in status_upper:
            cell_status.fill = fill_dudosa
            cell_status.font = font_dudosa
        elif "ERROR" in status_upper or "INVALIDA" in status_upper:
            cell_status.fill = fill_error
            cell_status.font = font_error

        # Cargar e insertar captura de pantalla
        screenshot_rel_path = data.get("screenshot", "")
        row_height = 20

        if screenshot_rel_path and screenshot_rel_path.startswith("/static/"):
            clean_rel_path = screenshot_rel_path.lstrip("/")
            abs_screenshot_path = os.path.join(app.root_path, clean_rel_path)

            if os.path.exists(abs_screenshot_path):
                try:
                    pil_img = PILImage.open(abs_screenshot_path)

                    # Miniatura
                    thumb_width = 180
                    w_percent = (thumb_width / float(pil_img.size[0]))
                    thumb_height = int((float(pil_img.size[1]) * float(w_percent)))

                    if thumb_height > 120:
                        thumb_height = 120
                        h_percent = (thumb_height / float(pil_img.size[1]))
                        thumb_width = int((float(pil_img.size[0]) * float(h_percent)))

                    pil_img = pil_img.resize(
                        (thumb_width, thumb_height),
                        PILImage.Resampling.LANCEWOOD if hasattr(PILImage, 'Resampling') else PILImage.ANTIALIAS
                    )

                    img_io = BytesIO()
                    pil_img.save(img_io, format="PNG")
                    img_io.seek(0)

                    img = OpenpyxlImage(img_io)

                    # Ajustar alto de fila
                    row_height = max(row_height, int(thumb_height * 0.75) + 15)
                    ws.row_dimensions[r_idx].height = row_height

                    # Añadir imagen
                    cell_loc = f"I{r_idx}"
                    ws.add_image(img, cell_loc)

                    # Texto / Enlace
                    cell_image = ws.cell(row=r_idx, column=9)
                    cell_image.value = "[Ver Imagen]"
                    cell_image.font = Font(name=font_family, size=9, color="0F6BFF", underline="single")
                    cell_image.alignment = Alignment(horizontal="center", vertical="bottom")
                    cell_image.hyperlink = abs_screenshot_path

                except Exception as e:
                    cell_image = ws.cell(row=r_idx, column=9, value=f"Error: {str(e)}")
                    cell_image.font = data_font
            else:
                ws.cell(row=r_idx, column=9, value="Sin archivo").font = data_font
        else:
            ws.cell(row=r_idx, column=9, value="No capturada").font = data_font

        ws.row_dimensions[r_idx].height = row_height

    # Ajustar ancho de columnas
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col_letter == "I":
            continue
        max_len = 0
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    ws.column_dimensions["I"].width = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="auditoria_final.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export-simple")
@login_required
def export_simple():
    results = get_state()["results"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen de Imagenes"

    # Estilos
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="0F2742", end_color="0F2742", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Colores por estado
    fill_valida = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_valida = Font(name=font_family, size=10, color="15803D", bold=True)

    fill_dudosa = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_dudosa = Font(name=font_family, size=10, color="A15C00", bold=True)

    fill_error = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_error = Font(name=font_family, size=10, color="C83232", bold=True)

    headers = [
        "URL Original",
        "Resultado",
        "Imagen"
    ]

    # Escribir cabeceras
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    ws.row_dimensions[1].height = 28
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 28

    # Recorrer resultados
    for r_idx, (url, data) in enumerate(results.items(), start=2):
        status = data.get("manual") or data.get("status", "")
        
        ws.cell(row=r_idx, column=1, value=url).alignment = align_left
        ws.cell(row=r_idx, column=2, value=status).alignment = align_center

        # Aplicar fuente
        ws.cell(row=r_idx, column=1).font = data_font
        ws.cell(row=r_idx, column=2).font = data_font

        # Estilo del estado
        cell_status = ws.cell(row=r_idx, column=2)
        status_upper = str(status).upper()
        if "VALIDA" in status_upper:
            cell_status.fill = fill_valida
            cell_status.font = font_valida
        elif "DUDOSA" in status_upper:
            cell_status.fill = fill_dudosa
            cell_status.font = font_dudosa
        elif "ERROR" in status_upper or "INVALIDA" in status_upper:
            cell_status.fill = fill_error
            cell_status.font = font_error

        # Cargar e insertar captura de pantalla
        screenshot_rel_path = data.get("screenshot", "")
        row_height = 20

        if screenshot_rel_path and screenshot_rel_path.startswith("/static/"):
            clean_rel_path = screenshot_rel_path.lstrip("/")
            abs_screenshot_path = os.path.join(app.root_path, clean_rel_path)

            if os.path.exists(abs_screenshot_path):
                try:
                    pil_img = PILImage.open(abs_screenshot_path)

                    # Miniatura
                    thumb_width = 180
                    w_percent = (thumb_width / float(pil_img.size[0]))
                    thumb_height = int((float(pil_img.size[1]) * float(w_percent)))

                    if thumb_height > 120:
                        thumb_height = 120
                        h_percent = (thumb_height / float(pil_img.size[1]))
                        thumb_width = int((float(pil_img.size[0]) * float(h_percent)))

                    pil_img = pil_img.resize(
                        (thumb_width, thumb_height),
                        PILImage.Resampling.LANCEWOOD if hasattr(PILImage, 'Resampling') else PILImage.ANTIALIAS
                    )

                    img_io = BytesIO()
                    pil_img.save(img_io, format="PNG")
                    img_io.seek(0)

                    img = OpenpyxlImage(img_io)

                    # Ajustar alto de fila
                    row_height = max(row_height, int(thumb_height * 0.75) + 15)
                    ws.row_dimensions[r_idx].height = row_height

                    # Añadir imagen
                    cell_loc = f"C{r_idx}"
                    ws.add_image(img, cell_loc)

                    # Texto / Enlace
                    cell_image = ws.cell(row=r_idx, column=3)
                    cell_image.value = "[Ver Imagen]"
                    cell_image.font = Font(name=font_family, size=9, color="0F6BFF", underline="single")
                    cell_image.alignment = Alignment(horizontal="center", vertical="bottom")
                    cell_image.hyperlink = abs_screenshot_path

                except Exception as e:
                    cell_image = ws.cell(row=r_idx, column=3, value=f"Error: {str(e)}")
                    cell_image.font = data_font
            else:
                ws.cell(row=r_idx, column=3, value="Sin archivo").font = data_font
        else:
            ws.cell(row=r_idx, column=3, value="No capturada").font = data_font

        ws.row_dimensions[r_idx].height = row_height

    # Ajustar ancho de columnas (excluyendo la columna C de la imagen)
    for col_idx in [1, 2]:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value:
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="auditoria_imagenes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
